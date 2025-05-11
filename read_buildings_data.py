import openpyxl
from datetime import datetime, timedelta
import csv
import os

def combine_and_export_data(input_file):
    # Load the workbook
    try:
        wb = openpyxl.load_workbook(input_file)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    # Get the sheets
    elec_sheet = wb.worksheets[1]  # Second sheet (electricity data)
    weather_sheet = wb.worksheets[2]  # Third sheet (weather data)

    # Prepare electricity data
    elec_data = []
    for row in elec_sheet.iter_rows(min_row=2, values_only=True):
        timestamp = row[0]
        if isinstance(timestamp, datetime):
            elec_data.append({
                'timestamp': timestamp,
                'values': row[1:11]  # Assuming columns B-K are Ap1-Ap10
            })

    # Prepare weather data
    weather_data = []
    weather_headers = []
    for col in range(1, weather_sheet.max_column + 1):
        header = weather_sheet.cell(row=1, column=col).value
        if header in ['Timestamps', 'air temperature',
                      'Atm pressure mm of mercury', 'Relative humidity (%)']:
            weather_headers.append(header)

    for row in weather_sheet.iter_rows(min_row=2, values_only=True):
        timestamp = row[0]
        if isinstance(timestamp, datetime):
            weather_values = []
            for header in weather_headers:
                if header != 'Timestamps':
                    index = weather_headers.index(header) + 1  # +1 because of 0-based index
                    weather_values.append(row[index])
            weather_data.append({
                'timestamp': timestamp,
                'values': weather_values
            })

    # Combine data
    combined = []
    for elec in elec_data:
        closest_weather = None
        min_diff = timedelta.max

        for weather in weather_data:
            time_diff = abs(weather['timestamp'] - elec['timestamp'])
            if time_diff < min_diff and time_diff <= timedelta(minutes=30):
                min_diff = time_diff
                closest_weather = weather

        if closest_weather:
            combined.append({
                'timestamp': elec['timestamp'],
                'elec_values': elec['values'],
                'weather_values': closest_weather['values']
            })

    # Define quarter groups
    quarters = [
        ('jan_mar', [1, 2, 3]),
        ('apr_jun', [4, 5, 6]),
        ('jul_sep', [7, 8, 9]),
        ('oct_dec', [10, 11, 12])
    ]

    # Write to CSV files
    for quarter_name, months in quarters:
        filename = f'{quarter_name}_combined.csv'
        try:
            with open(filename, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)

                # Write header
                writer.writerow(['Timestamps'] +
                                [f'Ap{i}' for i in range(1, 11)] +
                                weather_headers[1:])  # Exclude 'Timestamps' from weather headers

                # Write data
                for record in combined:
                    if record['timestamp'].month in months:
                        row = [record['timestamp'].strftime('%Y-%m-%d %H:%M:%S')] + \
                              list(record['elec_values']) + \
                              list(record['weather_values'])
                        writer.writerow(row)

            print(f'Created {filename}')
        except Exception as e:
            print(f"Error writing to {filename}: {e}")

# Usage
input_file = 'data/Buildings_aligned.xlsx'  # Replace with your file path
combine_and_export_data(input_file)
